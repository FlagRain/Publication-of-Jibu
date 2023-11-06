import json
import numpy as np
import openpyxl
import pandas as pd

place_dict = {}
df = pd.read_excel('./rm/明清书坊总数差值实际地点明.xlsx')
shufang = df.set_index('明书坊')['明'].to_dict()

#with open('./output/书坊数量-清.txt', mode='r') as f1:
#    shufang = json.load(f1)
with open('./output/版刻索引.txt', mode='r') as f2:
    dic = json.load(f2)

df2 = pd.read_excel('./rm/地名经纬度总表.xlsx')  # 替换 'your_excel_file.xlsx' 为您的 Excel 文件名或路径
place_list = df2.iloc[:, 0].tolist()
print(place_list)
for key in shufang.keys():
    a = str(list(df[df['明书坊'] == key]['实际地点'])[0])
    if a != 'nan':
        if a in place_dict.keys():
            place_dict[a] += int(shufang[key])
        else:
            place_dict[a] = int(shufang[key])
    else:
        for place in place_list:
            first_period_index = dic[key].find("。")
            result = dic[key][:first_period_index]
            if place in result:
                if place in place_dict.keys():
                    place_dict[place] += int(shufang[key])
                else:
                    place_dict[place] = int(shufang[key])
                break
res_list = []
for key, value in place_dict.items():
    for i in range(value):
        res_list.append(key)
workbook = openpyxl.Workbook()
sheet = workbook.active
for row, (key, value) in enumerate(place_dict.items(), start=1):
    sheet.cell(row=row, column=1, value=key)  # 写入键到第一列
    sheet.cell(row=row, column=2, value=value)  # 写入值到第二列
workbook.save("ming.xlsx")
workbook.close()
