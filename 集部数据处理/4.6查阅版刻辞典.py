import json
from fuzzywuzzy import fuzz
from zhconv import convert
import openpyxl


with open('./output/版刻索引.txt', mode='r') as f1:
    rm = json.load(f1)
dic = {}
keywords = ['局', '堂', '室', '书房', '书屋', '书院', '亭', '寺', '斋', '楼', '阁', '馆']
for keyword in keywords:
    dic.update({key: value for key, value in rm.items() if keyword in key})

with open('./output/中间.txt', mode='r') as f2:
    source = json.load(f2)
res_ming = {}
res_qing = {}
for lst in source.values():
    for name in lst:
        if name[0] == '明':
            name = convert(name, 'zh-hans')
            best_match = None
            best_ratio = 0
            for aim in dic.keys():
                ratio = fuzz.ratio(name, aim)
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = aim
            if best_ratio >= 60:
                if best_match in res_ming.keys():
                    res_ming[best_match] += 1
                else:
                    res_ming[best_match] = 1
        elif name[0] == '清':
            name = convert(name, 'zh-hans')
            best_match = None
            best_ratio = 0
            for aim in dic.keys():
                ratio = fuzz.ratio(name, aim)
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = aim
            if best_ratio >= 60:
                if best_match in res_qing.keys():
                    res_qing[best_match] += 1
                else:
                    res_qing[best_match] = 1
#with open('./output/书坊数量-明.txt', mode='w') as out:
#    json.dump(res_ming, out)
#with open('./output/书坊数量-清.txt', mode='w') as out2:
#    json.dump(res_qing, out2)
workbook = openpyxl.Workbook()
sheet = workbook.active
for row, (key, value) in enumerate(res_ming.items(), start=1):
    sheet.cell(row=row, column=1, value=key)  # 写入键到第一列
    sheet.cell(row=row, column=2, value=value)  # 写入值到第二列
workbook.save("明代书坊.xlsx")
workbook.close()
workbook = openpyxl.Workbook()
sheet = workbook.active
for row, (key, value) in enumerate(res_qing.items(), start=1):
    sheet.cell(row=row, column=1, value=key)  # 写入键到第一列
    sheet.cell(row=row, column=2, value=value)  # 写入值到第二列
workbook.save("清代书坊.xlsx")
workbook.close()